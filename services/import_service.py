from openpyxl import load_workbook
from database import SessionLocal, engine, Base
from models import (
    Session,
    Instrument,
    Strategy,
    ProbabilityLevel,
    MTFPhase,
    Trade,
)

EXCEL_PATH = r"D:\Web Development\Trading Journal\My Strategy Backtesting Template (2024) -  V1.2.xlsx"


def init_lookups(db):
    """Populate lookup tables from the Settings sheet."""
    wb = load_workbook(EXCEL_PATH, data_only=True)
    settings_sheet = wb["Settings"]

    sessions = []
    for row in range(4, 10):
        val = settings_sheet.cell(row=row, column=2).value
        if val:
            sessions.append(str(val).strip())

    instruments = []
    for row in range(4, 9):
        name = settings_sheet.cell(row=row, column=3).value
        if name:
            sl_buffer = settings_sheet.cell(row=row, column=10).value
            fixed_r_target = settings_sheet.cell(row=row, column=11).value
            instruments.append(
                {
                    "name": str(name).strip(),
                    "slug": str(name).strip().lower().replace(" ", "_"),
                    "sl_buffer": float(sl_buffer) if sl_buffer is not None else None,
                    "fixed_r_target": (
                        float(fixed_r_target)
                        if fixed_r_target is not None
                        else None
                    ),
                }
            )

    strategies = []
    for row in range(4, 12):
        val = settings_sheet.cell(row=row, column=4).value
        if val:
            strategies.append(str(val).strip())

    probabilities = []
    for row in range(4, 9):
        val = settings_sheet.cell(row=row, column=5).value
        if val:
            probabilities.append(str(val).strip())

    phases = []
    for row in range(4, 9):
        val = settings_sheet.cell(row=row, column=6).value
        if val:
            phases.append(str(val).strip())

    for name in sessions:
        db.add(Session(name=name, slug=name.lower().replace(" ", "_")))
    for inst in instruments:
        db.add(Instrument(**inst))
    for name in strategies:
        db.add(Strategy(name=name, slug=name.lower().replace(" ", "_")))
    for name in probabilities:
        db.add(
            ProbabilityLevel(name=name, slug=name.lower().replace(" ", "_"))
        )
    for name in phases:
        db.add(MTFPhase(name=name, slug=name.lower().replace(" ", "_")))

    db.commit()
    print(f"Inserted {len(sessions)} sessions, {len(instruments)} instruments, "
          f"{len(strategies)} strategies, {len(probabilities)} probabilities, "
          f"{len(phases)} phases.")


def get_lookup_maps(db):
    sessions = {r.name.lower().replace(" ", "_"): r.id for r in db.query(Session).all()}
    instruments = {r.name.lower().replace(" ", "_"): r.id for r in db.query(Instrument).all()}
    strategies = {r.name.lower().replace(" ", "_"): r.id for r in db.query(Strategy).all()}
    probabilities = {r.name.lower().replace(" ", "_"): r.id for r in db.query(ProbabilityLevel).all()}
    phases = {r.name.lower().replace(" ", "_"): r.id for r in db.query(MTFPhase).all()}
    return sessions, instruments, strategies, probabilities, phases


def safe_float(v):
    if v is None or v == "-" or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def safe_int(v):
    if v is None or v == "-" or v == "":
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def import_trades(db):
    wb = load_workbook(EXCEL_PATH, data_only=True)
    trades_sheet = wb["Trades"]

    sessions, instruments, strategies, probabilities, phases = get_lookup_maps(db)

    count = 0
    for row in range(3, trades_sheet.max_row + 1):
        date_val = trades_sheet.cell(row=row, column=1).value
        if date_val is None:
            continue

        session_name = trades_sheet.cell(row=row, column=3).value
        instrument_name = trades_sheet.cell(row=row, column=4).value
        strategy_name = trades_sheet.cell(row=row, column=5).value
        prob_name = trades_sheet.cell(row=row, column=6).value
        phase_name = trades_sheet.cell(row=row, column=7).value

        session_id = (
            sessions.get(str(session_name).strip().lower().replace(" ", "_"))
            if session_name else None
        )
        instrument_id = (
            instruments.get(str(instrument_name).strip().lower().replace(" ", "_"))
            if instrument_name else None
        )
        strategy_id = (
            strategies.get(str(strategy_name).strip().lower().replace(" ", "_"))
            if strategy_name else None
        )
        prob_id = (
            probabilities.get(str(prob_name).strip().lower().replace(" ", "_"))
            if prob_name else None
        )
        phase_id = (
            phases.get(str(phase_name).strip().lower().replace(" ", "_"))
            if phase_name else None
        )

        def float_val(cell):
            return safe_float(trades_sheet.cell(row=row, column=cell).value)

        def int_val(cell):
            return safe_int(trades_sheet.cell(row=row, column=cell).value)

        def str_val(cell):
            v = trades_sheet.cell(row=row, column=cell).value
            return str(v).strip() if v is not None else None

        trade = Trade(
            entry_date=date_val,
            entry_time=trades_sheet.cell(row=row, column=2).value,
            session_id=session_id,
            instrument_id=instrument_id,
            strategy_id=strategy_id,
            probability_level_id=prob_id,
            mtf_phase_id=phase_id,
            entry_news=str_val(8),
            management_news=str_val(9),
            entry_candle_size=float_val(10),
            mae_sl_buffer=float_val(11),
            mfe=float_val(12),
            max_r=float_val(13),
            fixed_r_target=float_val(14),
            screenshot_1=str_val(15),
            screenshot_2=str_val(16),
            screenshot_3=str_val(17),
            comments=str_val(18),
            day=int_val(19),
            day_of_week=int_val(20),
            month_number=int_val(21),
            month_name=str_val(22),
            year=int_val(23),
            quarter=int_val(24),
            total_r=float_val(25),
            is_win=bool(int_val(26)) if int_val(26) is not None else None,
            win_streak=int_val(27),
            is_loss=bool(int_val(28)) if int_val(28) is not None else None,
            loss_streak=int_val(29),
            peak_r=float_val(30),
            drawdown=float_val(31),
            max_drawdown=float_val(32),
        )
        db.add(trade)
        count += 1
        if count % 100 == 0:
            db.commit()
            print(f"Imported {count} trades...")

    db.commit()
    print(f"Total trades imported: {count}")


def run_import():
    Base.metadata.create_all(bind=engine)

    with SessionLocal() as db:
        if db.query(Instrument).first():
            print("Lookup tables already populated. Skipping init.")
        else:
            init_lookups(db)

        if db.query(Trade).first():
            print("Trades already imported. Skipping import.")
        else:
            import_trades(db)


if __name__ == "__main__":
    run_import()
